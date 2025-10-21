"""
Restore data from backup while preserving formatting.
Copies only cell VALUES, not formatting, from backup to current file.
"""

from openpyxl import load_workbook

print("=" * 60)
print("Restoring Data While Preserving Formatting")
print("=" * 60)

# Load both files
print("\nLoading files...")
backup_wb = load_workbook('backups/Weekly Tracker_2025-10-20.xlsx')
current_wb = load_workbook('Weekly Tracker.xlsx')

backup_ws = backup_wb['Weekly Gallons 25']
current_ws = current_wb['Weekly Gallons 25']

print("✓ Backup loaded")
print("✓ Current file loaded")

# Data columns to restore
columns = ['B', 'E', 'H', 'K']
column_names = ['Diesel', 'Regular', 'DEF', 'Total']

# Rows to restore (44-48)
rows_to_restore = range(44, 49)

print("\nRestoring data (values only, preserving formatting)...")
print("-" * 60)

for row in rows_to_restore:
    week_label = backup_ws[f'A{row}'].value
    print(f"\nRow {row} ({week_label}):")

    for col, name in zip(columns, column_names):
        cell_ref = f'{col}{row}'
        backup_value = backup_ws[cell_ref].value

        # Copy ONLY the value, not formatting
        current_ws[cell_ref].value = backup_value

        # Show what we copied
        if backup_value is not None:
            if isinstance(backup_value, (int, float)):
                print(f"  {name:8} ({col}): {backup_value:,.2f}")
            else:
                print(f"  {name:8} ({col}): {backup_value}")
        else:
            print(f"  {name:8} ({col}): (empty)")

# Save the updated file
print("\n" + "-" * 60)
print("Saving updated file...")
current_wb.save('Weekly Tracker.xlsx')
print("✓ File saved: Weekly Tracker.xlsx")

print("\n" + "=" * 60)
print("DATA RESTORATION COMPLETE")
print("=" * 60)
print("\nAll data restored while preserving cell formatting!")
print("You can now open the file to see the updated data.")
