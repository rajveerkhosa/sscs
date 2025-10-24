#!/usr/bin/env python3
"""
SSCS Weekly Report Automation
Collects data, generates Excel file, and emails the report.
"""

import os
import sys
import logging
import yaml
from datetime import datetime
from dotenv import load_dotenv

# Import our modules
from week_utils import get_week_params, format_sscs_datetime, get_last_year_week
from sscs_scraper import SSCSScraper
from fuel_aggregator import FuelAggregator
from storestats_scraper import StoreStatsScraper
from cstore_aggregator import CStoreAggregator
from excel_writer import ExcelReportWriter
from email_sender import EmailSender, format_email_body


def setup_logging(log_dir="logs"):
    """Setup logging to both file and console"""
    # Create logs directory
    os.makedirs(log_dir, exist_ok=True)

    # Create log file with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(log_dir, f'weekly_report_{timestamp}.log')

    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s: %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(sys.stdout)
        ]
    )

    logger = logging.getLogger(__name__)
    logger.info(f"Logging to: {log_file}")
    return logger


def get_all_departments(workbook_path):
    """Get all department numbers and names from Excel file"""
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    ws = wb.active

    departments = []
    dept_names = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        dept_cell = row[0]
        name_cell = row[1]

        if dept_cell.value and str(dept_cell.value).strip():
            try:
                dept_num = str(int(float(dept_cell.value)))
                departments.append(dept_num)
                dept_names[dept_num] = name_cell.value if name_cell.value else ""
            except (ValueError, TypeError):
                continue

    return departments, dept_names


def collect_all_data(logger):
    """
    Collect all SSCS data: Fuel, C-Store, and Department Sales.

    Returns:
        dict: All collected data or None if failed
    """
    load_dotenv()

    logger.info("=" * 70)
    logger.info("SSCS WEEKLY DATA COLLECTION - AUTOMATED")
    logger.info("=" * 70)

    # Load config
    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)

    # Get week parameters
    week_params = get_week_params()
    logger.info(f"\nWeek Label: {week_params['week_label']}")
    logger.info(f"Week Range: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")

    # Calculate last year's week
    ly_start_dt, ly_end_dt = get_last_year_week(week_params['end_datetime'])
    logger.info(f"Last Year Range: {ly_start_dt.date()} to {ly_end_dt.date()}")

    # Initialize scraper
    logger.info("\nInitializing SSCS scraper...")
    scraper = SSCSScraper(config, logger)
    headless = os.getenv('HEADLESS', 'true').lower() == 'true'
    scraper.start_browser(headless=headless)
    scraper.login()

    try:
        # Collect Fuel Data
        logger.info("\n" + "=" * 70)
        logger.info("COLLECTING FUEL DATA")
        logger.info("=" * 70)

        fuel_agg = FuelAggregator(config, scraper, logger)

        logger.info("\nThis Week:")
        fuel_data = fuel_agg.collect_all_gallons(week_params['start_date'], week_params['end_date'])

        logger.info("\nLast Year:")
        fuel_data_ly = fuel_agg.collect_all_gallons(
            format_sscs_datetime(ly_start_dt),
            format_sscs_datetime(ly_end_dt)
        )

        # Collect C-Store Data
        logger.info("\n" + "=" * 70)
        logger.info("COLLECTING C-STORE SALES DATA")
        logger.info("=" * 70)

        storestats = StoreStatsScraper(scraper, logger)
        cstore_agg = CStoreAggregator(config, storestats, logger)

        logger.info("\nThis Week:")
        cstore_data = cstore_agg.collect_all_sales(week_params['start_date'], week_params['end_date'])

        logger.info("\nLast Year:")
        cstore_data_ly = cstore_agg.collect_all_sales(
            format_sscs_datetime(ly_start_dt),
            format_sscs_datetime(ly_end_dt)
        )

        # Collect Department Sales (if file exists)
        dept_sales = {}
        dept_names = {}
        dept_file = "Weekly Department Sales.xlsx"

        if os.path.exists(dept_file):
            logger.info("\n" + "=" * 70)
            logger.info("COLLECTING DEPARTMENT SALES DATA")
            logger.info("=" * 70)

            departments, dept_names = get_all_departments(dept_file)
            logger.info(f"Found {len(departments)} departments to scrape")

            for i, dept in enumerate(departments, 1):
                logger.info(f"[{i}/{len(departments)}] Getting sales for Department {dept}...")
                try:
                    # Check browser connection
                    try:
                        scraper.driver.current_url
                    except Exception:
                        logger.error("Browser connection lost!")
                        break

                    sales = storestats.get_department_sales(
                        week_params['start_date'],
                        week_params['end_date'],
                        dept
                    )
                    dept_sales[dept] = sales

                except Exception as e:
                    logger.error(f"Failed to get sales for dept {dept}: {e}")
                    dept_sales[dept] = 0

            # Recheck zeros
            zero_depts = [d for d in departments if dept_sales.get(d, 0) == 0]
            if zero_depts:
                logger.info(f"\nRechecking {len(zero_depts)} departments with $0.00...")
                for dept in zero_depts[:5]:  # Limit rechecks to avoid long runtime
                    try:
                        sales = storestats.get_department_sales(
                            week_params['start_date'],
                            week_params['end_date'],
                            dept
                        )
                        if sales > 0:
                            logger.info(f"✓ Dept {dept} now shows ${sales:,.2f}")
                            dept_sales[dept] = sales
                    except:
                        pass

        # Close browser
        scraper.close()

        logger.info("\n" + "=" * 70)
        logger.info("✓ DATA COLLECTION COMPLETE")
        logger.info("=" * 70)

        return {
            'week_params': week_params,
            'fuel_data': fuel_data,
            'fuel_data_ly': fuel_data_ly,
            'cstore_data': cstore_data,
            'cstore_data_ly': cstore_data_ly,
            'dept_sales': dept_sales,
            'dept_names': dept_names,
            'config': config
        }

    except Exception as e:
        logger.error(f"Error during data collection: {e}")
        import traceback
        traceback.print_exc()
        scraper.close()
        return None


def generate_excel_report(data, logger):
    """
    Generate Excel report from collected data.

    Args:
        data (dict): Collected data
        logger: Logger instance

    Returns:
        str: Path to generated Excel file or None if failed
    """
    try:
        logger.info("\n" + "=" * 70)
        logger.info("GENERATING EXCEL REPORT")
        logger.info("=" * 70)

        week_params = data['week_params']
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = f"exports/SSCS_Weekly_Report_{week_params['week_label'].replace(' ', '_')}_{timestamp}.xlsx"

        # Create exports directory
        os.makedirs("exports", exist_ok=True)

        writer = ExcelReportWriter(excel_path)

        # Add sheets
        week_range = f"{week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}"

        writer.add_fuel_sheet(
            week_params['week_label'],
            week_range,
            data['fuel_data'],
            data['fuel_data_ly'],
            data['config']
        )

        writer.add_cstore_sheet(
            week_params['week_label'],
            week_range,
            data['cstore_data'],
            data['cstore_data_ly']
        )

        if data['dept_sales']:
            writer.add_department_sheet(
                week_params['dept_column_header'],
                week_range,
                data['dept_sales'],
                data['dept_names']
            )

        # Save
        saved_path = writer.save()
        logger.info(f"✓ Excel report saved: {saved_path}")

        return saved_path

    except Exception as e:
        logger.error(f"Failed to generate Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None


def send_email_report(data, excel_path, logger):
    """
    Send email report with Excel attachment.

    Args:
        data (dict): Collected data
        excel_path (str): Path to Excel file
        logger: Logger instance

    Returns:
        bool: True if sent successfully
    """
    try:
        logger.info("\n" + "=" * 70)
        logger.info("SENDING EMAIL REPORT")
        logger.info("=" * 70)

        recipient = os.getenv('EMAIL_TO')
        if not recipient:
            logger.warning("EMAIL_TO not set in .env file - skipping email")
            return False

        week_params = data['week_params']
        week_range = f"{week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}"

        # Format email body
        body = format_email_body(
            week_params['week_label'],
            week_range,
            data['fuel_data'],
            data['fuel_data_ly'],
            data['cstore_data'],
            data['cstore_data_ly'],
            dept_count=len(data['dept_sales']),
            dept_sales=data['dept_sales'],
            dept_names=data['dept_names']
        )

        # Send email
        sender = EmailSender(logger=logger)
        subject = f"SSCS Weekly Report - {week_params['week_label']}"

        success = sender.send_weekly_report(
            recipient=recipient,
            subject=subject,
            body_text=body,
            excel_path=excel_path
        )

        return success

    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main entry point for weekly report automation"""
    logger = setup_logging()

    logger.info("\n" + "=" * 70)
    logger.info("SSCS WEEKLY REPORT AUTOMATION")
    logger.info("=" * 70)
    logger.info(f"Started: {datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')}")

    # Step 1: Collect data
    data = collect_all_data(logger)
    if not data:
        logger.error("\n✗ DATA COLLECTION FAILED")
        sys.exit(1)

    # Step 2: Generate Excel (DISABLED - Future implementation)
    # excel_path = generate_excel_report(data, logger)
    # if not excel_path:
    #     logger.error("\n✗ EXCEL GENERATION FAILED")
    #     sys.exit(1)
    excel_path = None
    logger.info("\n" + "=" * 70)
    logger.info("Excel generation disabled (future feature)")
    logger.info("=" * 70)

    # Step 3: Send email (without Excel attachment)
    email_sent = send_email_report(data, excel_path, logger)

    # Final summary
    logger.info("\n" + "=" * 70)
    logger.info("WEEKLY REPORT AUTOMATION - SUMMARY")
    logger.info("=" * 70)
    logger.info(f"✓ Data Collection: SUCCESS")
    logger.info(f"- Excel Generation: DISABLED (future feature)")
    logger.info(f"{'✓' if email_sent else '✗'} Email Delivery: {'SUCCESS' if email_sent else 'FAILED/SKIPPED'}")
    logger.info(f"Completed: {datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')}")
    logger.info("=" * 70)

    if not email_sent:
        logger.warning("\nEmail was not sent. Check EMAIL_TO, EMAIL_USER, and EMAIL_PASSWORD in .env")
    else:
        logger.info("\n✓ Weekly report email sent successfully (text only, no Excel attachment)")

    return 0 if email_sent else 1


if __name__ == '__main__':
    sys.exit(main())
