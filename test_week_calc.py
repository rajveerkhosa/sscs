#!/usr/bin/env python3
"""
Quick test script to verify week calculation and configuration.
Run this before running the full main.py script.
"""

import os
from datetime import datetime
from week_utils import get_week_params, get_ordinal_suffix, format_week_label


def test_ordinal_suffix():
    """Test ordinal suffix generation"""
    print("Testing ordinal suffix generation...")
    test_cases = {
        1: '1st', 2: '2nd', 3: '3rd', 4: '4th', 5: '5th',
        11: '11th', 12: '12th', 13: '13th',
        21: '21st', 22: '22nd', 23: '23rd',
        31: '31st'
    }

    for day, expected in test_cases.items():
        result = f"{day}{get_ordinal_suffix(day)}"
        status = "✓" if result == expected else "✗"
        print(f"  {status} Day {day}: {result} (expected {expected})")


def test_week_label():
    """Test week label formatting"""
    print("\nTesting week label formatting...")
    test_dates = [
        datetime(2025, 1, 15),
        datetime(2025, 3, 3),
        datetime(2025, 12, 21),
    ]

    for date in test_dates:
        label = format_week_label(date)
        print(f"  {date.strftime('%Y-%m-%d')} → {label}")


def test_week_params():
    """Test week parameter calculation"""
    print("\nTesting week parameter calculation...")
    params = get_week_params()

    print(f"\nToday: {datetime.now().strftime('%Y-%m-%d %A')}")
    print(f"\nLast Full Week:")
    print(f"  Week Ending: {params['week_label']}")
    print(f"  Start: {params['start_datetime'].strftime('%Y-%m-%d %A %H:%M:%S')}")
    print(f"  End:   {params['end_datetime'].strftime('%Y-%m-%d %A %H:%M:%S')}")
    print(f"\nSSCS Format:")
    print(f"  Start: {params['start_date']}")
    print(f"  End:   {params['end_date']}")

    # Verify it's actually a Monday to Sunday
    assert params['start_datetime'].weekday() == 0, "Start should be Monday"
    assert params['end_datetime'].weekday() == 6, "End should be Sunday"
    print("\n✓ Week validation passed (Monday to Sunday)")


def test_config():
    """Test configuration file"""
    print("\n" + "=" * 60)
    print("Testing Configuration")
    print("=" * 60)

    if not os.path.exists('config.yaml'):
        print("✗ config.yaml not found!")
        return False

    print("✓ config.yaml found")

    try:
        import yaml
        with open('config.yaml', 'r') as f:
            config = yaml.safe_load(f)
        print("✓ config.yaml is valid YAML")

        # Check key sections
        required_keys = ['sscs', 'week', 'fuel', 'excel']
        for key in required_keys:
            if key in config:
                print(f"✓ Section '{key}' found")
            else:
                print(f"✗ Section '{key}' missing!")
                return False

        return True
    except Exception as e:
        print(f"✗ Error loading config: {e}")
        return False


def test_env():
    """Test environment file"""
    print("\n" + "=" * 60)
    print("Testing Environment")
    print("=" * 60)

    if not os.path.exists('.env'):
        print("✗ .env file not found!")
        print("  Copy .env.template to .env and add your credentials")
        return False

    print("✓ .env file found")

    from dotenv import load_dotenv
    load_dotenv()

    user = os.getenv('SSCS_USER')
    password = os.getenv('SSCS_PASS')

    if user:
        print(f"✓ SSCS_USER is set ('{user[:3]}...')")
    else:
        print("✗ SSCS_USER not set in .env")
        return False

    if password:
        print(f"✓ SSCS_PASS is set (length: {len(password)})")
    else:
        print("✗ SSCS_PASS not set in .env")
        return False

    return True


def test_directories():
    """Test required directories"""
    print("\n" + "=" * 60)
    print("Testing Directories")
    print("=" * 60)

    required_dirs = ['exports', 'backups', 'logs']
    all_exist = True

    for dir_name in required_dirs:
        if os.path.exists(dir_name):
            print(f"✓ {dir_name}/ exists")
        else:
            print(f"✗ {dir_name}/ missing (will be created on first run)")
            all_exist = False

    return all_exist


def test_excel_file():
    """Test Excel tracker file"""
    print("\n" + "=" * 60)
    print("Testing Excel File")
    print("=" * 60)

    tracker_file = 'Weekly Tracker.xlsx'

    if not os.path.exists(tracker_file):
        print(f"✗ {tracker_file} not found in current directory!")
        print("  Place your tracker file here before running main.py")
        return False

    print(f"✓ {tracker_file} found")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(tracker_file, read_only=True)
        print(f"✓ Excel file is valid (sheets: {', '.join(wb.sheetnames)})")
        wb.close()
        return True
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        return False


def main():
    """Run all tests"""
    print("=" * 60)
    print("SSCS Tracker - Configuration Test")
    print("=" * 60)

    # Test week calculations
    test_ordinal_suffix()
    test_week_label()
    test_week_params()

    # Test configuration
    config_ok = test_config()
    env_ok = test_env()
    dirs_ok = test_directories()
    excel_ok = test_excel_file()

    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)

    all_ok = config_ok and env_ok and excel_ok

    if all_ok:
        print("✓ All checks passed!")
        print("\nYou're ready to run: python main.py")
    else:
        print("✗ Some checks failed")
        print("\nFix the issues above before running main.py")

    if not dirs_ok:
        print("\nNote: Missing directories will be created automatically")

    print("=" * 60)


if __name__ == '__main__':
    main()
