#!/usr/bin/env python3
"""
Get all weekly data - Fuel + C-Store for ANY week.
Displays everything in the terminal for manual entry.
"""

import os
import sys
import logging
import yaml
from datetime import datetime, timedelta
from dotenv import load_dotenv

from week_utils import get_week_params, format_sscs_datetime, get_last_year_week, format_week_label
from sscs_scraper import SSCSScraper
from fuel_aggregator import FuelAggregator
from storestats_scraper import StoreStatsScraper
from cstore_aggregator import CStoreAggregator
from openpyxl import load_workbook


def get_week_data_for_date(week_ending_date):
    """
    Calculate week parameters for a specific week ending date.

    Args:
        week_ending_date (datetime): The Sunday ending the week

    Returns:
        dict: Week parameters including dates and labels
    """
    # Ensure it's a Sunday
    if week_ending_date.weekday() != 6:
        # Adjust to the nearest Sunday
        days_to_sunday = (6 - week_ending_date.weekday()) % 7
        if days_to_sunday == 0:
            days_to_sunday = 7
        week_ending_date = week_ending_date + timedelta(days=days_to_sunday)

    # Calculate Monday to Sunday
    week_ending_sunday = week_ending_date.replace(hour=23, minute=59, second=59, microsecond=0)
    week_starting_monday = (week_ending_date - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)

    # Excel label is the Monday AFTER the week ends
    excel_label_date = week_ending_sunday + timedelta(days=1)
    week_label = format_week_label(excel_label_date)

    # Department sales header format: "20th Oct" (day first)
    day = excel_label_date.day
    suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10 if day % 100 not in [11, 12, 13] else 0, 'th')
    month_abbrev = excel_label_date.strftime('%b')
    dept_column_header = f"{day}{suffix} {month_abbrev}"

    # Get last year week
    ly_start_dt, ly_end_dt = get_last_year_week(week_ending_sunday)

    return {
        'week_label': week_label,
        'dept_column_header': dept_column_header,
        'start_datetime': week_starting_monday,
        'end_datetime': week_ending_sunday,
        'start_date': format_sscs_datetime(week_starting_monday),
        'end_date': format_sscs_datetime(week_ending_sunday),
        'ly_start_datetime': ly_start_dt,
        'ly_end_datetime': ly_end_dt,
        'ly_start_date': format_sscs_datetime(ly_start_dt),
        'ly_end_date': format_sscs_datetime(ly_end_dt)
    }


def get_all_departments(workbook_path):
    """Get all department numbers from Weekly Department Sales.xlsx"""
    if not os.path.exists(workbook_path):
        return []

    wb = load_workbook(workbook_path)
    ws = wb.active

    departments = []
    dept_names = {}

    # Start from row 2 (skip header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        dept_cell = row[0]
        name_cell = row[1]

        if dept_cell.value and str(dept_cell.value).strip():
            try:
                dept_num = str(int(float(dept_cell.value)))
                departments.append(dept_num)
                dept_names[dept_num] = name_cell.value if name_cell.value else ""
            except (ValueError, TypeError):
                continue

    return departments, dept_names


def main():
    """Get all data and display it"""

    load_dotenv()

    logging.basicConfig(
        level=logging.INFO,
        format='%(levelname)s: %(message)s'
    )

    logger = logging.getLogger(__name__)

    print("=" * 70)
    print("SSCS WEEKLY DATA COLLECTION")
    print("=" * 70)

    # Load config
    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)

    # Ask user for date or use current week
    print("\nOptions:")
    print("1. Current week (automatic)")
    print("2. Specific week (enter a date)")
    choice = input("Select option [1/2]: ").strip()

    if choice == "2":
        date_input = input("Enter week ending Sunday (YYYY-MM-DD) or any date in the week: ").strip()
        try:
            input_date = datetime.strptime(date_input, "%Y-%m-%d")
            week_params = get_week_data_for_date(input_date)
        except ValueError:
            print("Invalid date format. Using current week instead.")
            base_params = get_week_params()
            week_params = get_week_data_for_date(base_params['end_datetime'])
    else:
        base_params = get_week_params()
        week_params = get_week_data_for_date(base_params['end_datetime'])

    print(f"\nExcel Row Label: {week_params['week_label']}")
    print(f"This Week Data: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")
    print(f"Last Year Data: {week_params['ly_start_datetime'].date()} to {week_params['ly_end_datetime'].date()}")

    # Initialize scraper
    logger.info("\nInitializing SSCS scraper...")
    scraper = SSCSScraper(config, logger)

    # Start browser (visible so you can see progress)
    scraper.start_browser(headless=False)
    scraper.login()

    try:
        # Initialize aggregators
        fuel_agg = FuelAggregator(config, scraper, logger)
        storestats = StoreStatsScraper(scraper, logger)
        cstore_agg = CStoreAggregator(config, storestats, logger)

        print("\n" + "=" * 70)
        print("COLLECTING FUEL DATA")
        print("=" * 70)

        # Collect fuel data for THIS WEEK
        logger.info("\nCollecting THIS WEEK fuel data...")
        fuel_data = fuel_agg.collect_all_gallons(
            week_params['start_date'],
            week_params['end_date']
        )

        # Collect fuel data for LAST YEAR
        logger.info("\nCollecting LAST YEAR fuel data...")
        fuel_data_ly = fuel_agg.collect_all_gallons(
            week_params['ly_start_date'],
            week_params['ly_end_date']
        )

        print("\n" + "=" * 70)
        print("COLLECTING C-STORE DATA")
        print("=" * 70)

        # Collect C-Store data for THIS WEEK
        logger.info("\nCollecting THIS WEEK C-Store data...")
        cstore_data = cstore_agg.collect_all_sales(
            week_params['start_date'],
            week_params['end_date']
        )

        # Collect C-Store data for LAST YEAR
        logger.info("\nCollecting LAST YEAR C-Store data...")
        cstore_data_ly = cstore_agg.collect_all_sales(
            week_params['ly_start_date'],
            week_params['ly_end_date']
        )

        # Display all data in a nice format
        print("\n" + "=" * 70)
        print("WEEKLY GALLONS 25 - FUEL DATA")
        print("=" * 70)
        print(f"Excel Row Label: {week_params['week_label']}")
        print(f"Week Data Range: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")
        print("=" * 70)
        print(f"{'':15} {'THIS WEEK':>15} {'LAST YEAR':>15} {'CHANGE':>15}")
        print("-" * 70)

        diesel_change = fuel_data['diesel_gal'] - fuel_data_ly['diesel_gal']
        regular_change = fuel_data['regular_gal'] - fuel_data_ly['regular_gal']
        def_change = fuel_data['def_gal'] - fuel_data_ly['def_gal']
        total_change = fuel_data['total_gal'] - fuel_data_ly['total_gal']

        # Summary only - no breakdown
        print(f"{'Diesel (B/C):':<15} {fuel_data['diesel_gal']:>15,.2f} {fuel_data_ly['diesel_gal']:>15,.2f} {diesel_change:>+15,.2f}")
        print(f"{'Regular (E/F):':<15} {fuel_data['regular_gal']:>15,.2f} {fuel_data_ly['regular_gal']:>15,.2f} {regular_change:>+15,.2f}")
        print(f"{'DEF (H/I):':<15} {fuel_data['def_gal']:>15,.2f} {fuel_data_ly['def_gal']:>15,.2f} {def_change:>+15,.2f}")
        print("-" * 70)
        print(f"{'TOTAL (K/L):':<15} {fuel_data['total_gal']:>15,.2f} {fuel_data_ly['total_gal']:>15,.2f} {total_change:>+15,.2f}")
        print("=" * 70)

        print("\n" + "=" * 70)
        print("C STORE SALES 25 - SALES DATA")
        print("=" * 70)
        print(f"Excel Row Label: {week_params['week_label']}")
        print(f"Week Data Range: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")
        print("=" * 70)
        print(f"{'':20} {'THIS WEEK':>15} {'LAST YEAR':>15} {'CHANGE':>15}")
        print("-" * 70)

        total_change = cstore_data['total_cstore_sales'] - cstore_data_ly['total_cstore_sales']
        lottery_change = cstore_data['lottery_sales'] - cstore_data_ly['lottery_sales']
        scale_change = cstore_data['scale_sales'] - cstore_data_ly['scale_sales']
        other_change = cstore_data['other_sales'] - cstore_data_ly['other_sales']

        # Summary only - no breakdown
        print(f"{'Total C-Store (B/C):':<20} ${cstore_data['total_cstore_sales']:>14,.2f} ${cstore_data_ly['total_cstore_sales']:>14,.2f} ${total_change:>+14,.2f}")
        print(f"{'Lottery (E/F):':<20} ${cstore_data['lottery_sales']:>14,.2f} ${cstore_data_ly['lottery_sales']:>14,.2f} ${lottery_change:>+14,.2f}")
        print(f"{'Scale (H/I):':<20} ${cstore_data['scale_sales']:>14,.2f} ${cstore_data_ly['scale_sales']:>14,.2f} ${scale_change:>+14,.2f}")
        print(f"{'Other Sales (K/L):':<20} ${cstore_data['other_sales']:>14,.2f} ${cstore_data_ly['other_sales']:>14,.2f} ${other_change:>+14,.2f}")
        print("=" * 70)

        # Department Sales
        dept_file = "Weekly Department Sales.xlsx"
        if os.path.exists(dept_file):
            logger.info(f"\nFound {dept_file}, collecting department sales...")

            departments, dept_names = get_all_departments(dept_file)

            if departments:
                print("\n" + "=" * 70)
                print(f"WEEKLY DEPARTMENT SALES - {week_params['dept_column_header']}")
                print("=" * 70)
                print(f"Week: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")
                print(f"Column Header: {week_params['dept_column_header']}")
                print("=" * 70)

                dept_sales = {}
                failed_depts = []

                for i, dept in enumerate(departments, 1):
                    logger.info(f"[{i}/{len(departments)}] Getting sales for Department {dept}...")
                    try:
                        # Check if browser is still connected
                        try:
                            scraper.driver.current_url
                        except Exception:
                            logger.error("Browser connection lost! Stopping...")
                            print("\n⚠ Browser connection lost during scraping!")
                            print(f"Successfully scraped {i-1} out of {len(departments)} departments")
                            break

                        sales = storestats.get_department_sales(
                            week_params['start_date'],
                            week_params['end_date'],
                            dept
                        )
                        dept_sales[dept] = sales
                    except Exception as e:
                        logger.error(f"Failed to get sales for dept {dept}: {e}")
                        dept_sales[dept] = 0
                        failed_depts.append(dept)

                # Second pass: Re-check departments that returned 0.00
                zero_depts = [dept for dept in departments if dept_sales.get(dept, 0) == 0 and dept not in failed_depts]

                if zero_depts:
                    print("\n" + "=" * 70)
                    print(f"RECHECKING {len(zero_depts)} DEPARTMENTS WITH $0.00")
                    print("=" * 70)

                    for i, dept in enumerate(zero_depts, 1):
                        logger.info(f"[Recheck {i}/{len(zero_depts)}] Retrying Department {dept}...")
                        try:
                            # Check browser connection
                            try:
                                scraper.driver.current_url
                            except Exception:
                                logger.error("Browser connection lost during recheck!")
                                break

                            sales = storestats.get_department_sales(
                                week_params['start_date'],
                                week_params['end_date'],
                                dept
                            )

                            if sales > 0:
                                logger.info(f"✓ Dept {dept} now shows ${sales:,.2f} (was $0.00)")
                                dept_sales[dept] = sales
                            else:
                                logger.info(f"Dept {dept} still $0.00 after recheck")

                        except Exception as e:
                            logger.error(f"Failed recheck for dept {dept}: {e}")

                # Display results
                print(f"{'Dept #':<10} {'Department Name':<30} {'Sales':<15} {'Status':<10}")
                print("-" * 80)

                total_dept_sales = 0
                for dept in departments:
                    if dept in dept_sales:
                        sales = dept_sales[dept]
                        total_dept_sales += sales
                        dept_name = dept_names.get(dept, "")
                        status = "✓" if dept not in failed_depts else "✗ FAILED"
                        print(f"{dept:<10} {dept_name:<30} ${sales:>13,.2f} {status:<10}")
                    else:
                        dept_name = dept_names.get(dept, "")
                        print(f"{dept:<10} {dept_name:<30} {'NOT SCRAPED':>15} {'⚠ SKIPPED':<10}")

                print("-" * 80)
                print(f"{'TOTAL':<10} {'':30} ${total_dept_sales:>13,.2f}")
                print("=" * 70)

                # Summary
                if failed_depts:
                    print(f"\n⚠ WARNING: Failed to get data for {len(failed_depts)} department(s): {', '.join(failed_depts)}")
                    print("These departments have $0.00 - you may need to manually check them.")

                scraped_count = len([d for d in departments if d in dept_sales and d not in failed_depts])
                print(f"\n✓ Successfully scraped {scraped_count} out of {len(departments)} departments")
            else:
                logger.info(f"No departments found in {dept_file}")
        else:
            logger.info(f"{dept_file} not found, skipping department sales")

        # Close browser after all scraping is complete
        scraper.close()

        # Copy-paste section for ALL data
        print("\n" + "=" * 70)
        print("COPY-PASTE VALUES FOR EXCEL - ALL DATA")
        print("=" * 70)

        print("\n--- WEEKLY GALLONS 25 ---")
        print(f"Excel Row Label: {week_params['week_label']}")
        print(f"Week: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")
        print("\nTHIS WEEK (Columns B, E, H, K):")
        print(f"{fuel_data['diesel_gal']:.2f}")
        print(f"{fuel_data['regular_gal']:.2f}")
        print(f"{fuel_data['def_gal']:.2f}")
        print(f"{fuel_data['total_gal']:.2f}")

        print("\nLAST YEAR (Columns C, F, I, L):")
        print(f"{fuel_data_ly['diesel_gal']:.2f}")
        print(f"{fuel_data_ly['regular_gal']:.2f}")
        print(f"{fuel_data_ly['def_gal']:.2f}")
        print(f"{fuel_data_ly['total_gal']:.2f}")

        # Add detailed fuel breakdown
        print("\n" + "-" * 70)
        print("FUEL DATA BREAKDOWN BY PREFIX")
        print("-" * 70)

        diesel_prefixes = config['fuel']['diesel_prefixes']
        regular_prefixes = config['fuel']['regular_prefixes']
        def_prefixes = config['fuel']['def_prefixes']

        print(f"\n{'Prefix':<10} {'Type':<15} {'THIS WEEK':>15} {'LAST YEAR':>15} {'CHANGE':>15}")
        print("-" * 70)

        for prefix in diesel_prefixes:
            tw_val = fuel_data['prefix_details'].get(prefix, 0)
            ly_val = fuel_data_ly['prefix_details'].get(prefix, 0)
            change = tw_val - ly_val
            print(f"{prefix:<10} {'Diesel':<15} {tw_val:>15,.2f} {ly_val:>15,.2f} {change:>+15,.2f}")

        for prefix in regular_prefixes:
            tw_val = fuel_data['prefix_details'].get(prefix, 0)
            ly_val = fuel_data_ly['prefix_details'].get(prefix, 0)
            change = tw_val - ly_val
            print(f"{prefix:<10} {'Regular':<15} {tw_val:>15,.2f} {ly_val:>15,.2f} {change:>+15,.2f}")

        for prefix in def_prefixes:
            tw_val = fuel_data['prefix_details'].get(prefix, 0)
            ly_val = fuel_data_ly['prefix_details'].get(prefix, 0)
            change = tw_val - ly_val
            print(f"{prefix:<10} {'DEF':<15} {tw_val:>15,.2f} {ly_val:>15,.2f} {change:>+15,.2f}")

        print("\n--- C STORE SALES 25 ---")
        print(f"Excel Row Label: {week_params['week_label']}")
        print(f"Week: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")
        print("\nTHIS WEEK (Columns B, E, H, K):")
        print(f"{cstore_data['total_cstore_sales']:.2f}")
        print(f"{cstore_data['lottery_sales']:.2f}")
        print(f"{cstore_data['scale_sales']:.2f}")
        print(f"{cstore_data['other_sales']:.2f}")

        print("\nLAST YEAR (Columns C, F, I, L):")
        print(f"{cstore_data_ly['total_cstore_sales']:.2f}")
        print(f"{cstore_data_ly['lottery_sales']:.2f}")
        print(f"{cstore_data_ly['scale_sales']:.2f}")
        print(f"{cstore_data_ly['other_sales']:.2f}")

        # Add C-Store breakdown
        print("\n" + "-" * 70)
        print("C STORE SALES BREAKDOWN BY DEPARTMENT")
        print("-" * 70)
        print(f"\n{'Department':<15} {'Category':<15} {'THIS WEEK':>15} {'LAST YEAR':>15} {'CHANGE':>15}")
        print("-" * 70)

        for dept_num in ['27', '43', '72']:
            tw_val = cstore_data.get(f'dept_{dept_num}', 0)
            ly_val = cstore_data_ly.get(f'dept_{dept_num}', 0)
            change = tw_val - ly_val
            print(f"{'Dept ' + dept_num:<15} {'Lottery':<15} ${tw_val:>14,.2f} ${ly_val:>14,.2f} ${change:>+14,.2f}")

        tw_val = cstore_data.get('dept_88', 0)
        ly_val = cstore_data_ly.get('dept_88', 0)
        change = tw_val - ly_val
        print(f"{'Dept 88':<15} {'Scale':<15} ${tw_val:>14,.2f} ${ly_val:>14,.2f} ${change:>+14,.2f}")

        # Department sales if available
        if os.path.exists(dept_file):
            departments, dept_names = get_all_departments(dept_file)
            if departments and 'dept_sales' in locals():
                print(f"\n--- WEEKLY DEPARTMENT SALES ({week_params['dept_column_header']}) ---")
                print(f"Column Header: {week_params['dept_column_header']}")
                print(f"Week: {week_params['start_datetime'].date()} to {week_params['end_datetime'].date()}")
                print("\nDepartment Sales (paste into new column):")
                total_dept_sales = 0
                for dept in departments:
                    sales = dept_sales.get(dept, 0)
                    total_dept_sales += sales
                    print(f"{sales:.2f}")
                print(f"\nTotal: {total_dept_sales:.2f}")

        print("\n" + "=" * 70)

        print("\n✓ All data collected successfully!")
        print("\nYou can now manually enter these values into Excel.")

    except Exception as e:
        logger.error(f"Error collecting data: {e}")
        import traceback
        traceback.print_exc()
        scraper.close()
        sys.exit(1)


if __name__ == '__main__':
    main()
