"""
Fix Excel formulas:
1. Extend percentage formulas (D, G, J, M) to all data rows (44-48)
2. Update Total row (49) SUM formulas to include all data rows (4-48)
"""

from openpyxl import load_workbook

print("=" * 70)
print("Fixing Excel Formulas")
print("=" * 70)

# Load workbook
wb = load_workbook('Weekly Tracker.xlsx')
ws = wb['Weekly Gallons 25']

print("\n1. Extending percentage formulas to rows 44-48...")
print("-" * 70)

# Percentage columns: D, G, J, M
# Formula pattern: D=B/C, G=E/F, J=H/I, M=K/L

percentage_columns = {
    'D': ('B', 'C'),  # Diesel %
    'G': ('E', 'F'),  # Regular %
    'J': ('H', 'I'),  # DEF %
    'M': ('K', 'L')   # Total %
}

# Apply to rows 44-48
for row in range(44, 49):
    week_label = ws[f'A{row}'].value
    print(f"\nRow {row} ({week_label}):")

    for pct_col, (numerator_col, denominator_col) in percentage_columns.items():
        formula = f"={numerator_col}{row}/{denominator_col}{row}"
        ws[f'{pct_col}{row}'].value = formula
        print(f"  {pct_col}{row}: {formula}")

print("\n" + "=" * 70)
print("2. Updating Total row (49) SUM formulas...")
print("-" * 70)

# Update SUM formulas in row 49 to sum from row 4 to row 48
sum_columns = ['B', 'E', 'H', 'K']  # Diesel, Regular, DEF, Total

for col in sum_columns:
    formula = f"=sum({col}4:{col}48)"
    ws[f'{col}49'].value = formula
    print(f"  {col}49: {formula}")

# Percentage formulas in row 49 (already correct, but verify)
print("\nRow 49 percentage formulas (unchanged):")
for pct_col, (numerator_col, denominator_col) in percentage_columns.items():
    current_formula = ws[f'{pct_col}49'].value
    print(f"  {pct_col}49: {current_formula}")

print("\n" + "=" * 70)
print("Saving changes...")
wb.save('Weekly Tracker.xlsx')
print("✓ File saved: Weekly Tracker.xlsx")

print("\n" + "=" * 70)
print("FORMULAS FIXED SUCCESSFULLY")
print("=" * 70)
print("\n✓ Percentage formulas extended to all data rows (44-48)")
print("✓ Total row SUM formulas updated to include rows 4-48")
print("\nYou can now open the file to verify the calculations!")
