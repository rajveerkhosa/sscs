"""
Excel updater for Weekly Tracker.xlsx
Handles backup, row insertion, formatting preservation, and rolling window logic.
"""

import os
import shutil
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy


class ExcelUpdater:
    """Updates the Weekly Tracker Excel file"""

    def __init__(self, config, logger=None):
        """
        Initialize updater.

        Args:
            config (dict): Configuration dictionary
            logger (logging.Logger, optional): Logger instance
        """
        self.config = config
        self.logger = logger or logging.getLogger(__name__)

        self.workbook_path = config['excel']['workbook']
        self.backup_dir = config['excel']['backup_dir']

        if not os.path.exists(self.workbook_path):
            raise FileNotFoundError(f"Tracker file not found: {self.workbook_path}")

    def backup_workbook(self):
        """
        Create a dated backup of the workbook.

        Returns:
            str: Path to backup file
        """
        today = datetime.now().strftime('%Y-%m-%d')
        filename = os.path.basename(self.workbook_path)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_{today}{ext}"
        backup_path = os.path.join(self.backup_dir, backup_filename)

        shutil.copy2(self.workbook_path, backup_path)
        self.logger.info(f"Backup created: {backup_path}")

        return backup_path

    def update_tracker(self, week_label, fuel_data, cstore_data=None):
        """
        Update the tracker with new week data.

        Args:
            week_label (str): Week ending label (e.g., '22nd Sep')
            fuel_data (dict): Fuel gallons data (diesel_gal, regular_gal, def_gal, total_gal)
            cstore_data (dict, optional): C-Store sales data (cstore_sales)
        """
        self.logger.info(f"Updating tracker for week: {week_label}")

        # Backup first
        self.backup_workbook()

        # Load workbook
        wb = load_workbook(self.workbook_path)

        # Update each enabled sheet
        for sheet_config in self.config['excel']['sheets']:
            sheet_name = sheet_config['name']

            # Skip if sheet is disabled
            if not sheet_config.get('enabled', True):
                self.logger.info(f"Skipping disabled sheet: {sheet_name}")
                continue

            # Skip C Store sheet if no data provided
            if 'C Store' in sheet_name and cstore_data is None:
                self.logger.info(f"Skipping {sheet_name} (no c-store data)")
                continue

            if sheet_name not in wb.sheetnames:
                self.logger.error(f"Sheet not found: {sheet_name}")
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

            self.logger.info(f"Updating sheet: {sheet_name}")
            ws = wb[sheet_name]

            # Update the sheet
            self._update_sheet(ws, sheet_config, week_label, fuel_data, cstore_data)

        # Save workbook
        wb.save(self.workbook_path)
        self.logger.info(f"Tracker saved: {self.workbook_path}")

    def _update_sheet(self, ws, sheet_config, week_label, fuel_data, cstore_data):
        """
        Update a single sheet with new week data.

        Args:
            ws: openpyxl worksheet
            sheet_config (dict): Sheet configuration
            week_label (str): Week ending label
            fuel_data (dict): Fuel data
            cstore_data (dict): C-Store data
        """
        week_col = sheet_config['week_col']
        total_row_label = sheet_config['total_row_label']
        thisweek_columns = sheet_config['thisweek_columns']

        # Find the Total row
        total_row_num = self._find_total_row(ws, week_col, total_row_label)
        if total_row_num is None:
            raise ValueError(f"Total row not found in sheet {ws.title}")

        self.logger.info(f"Total row found at row {total_row_num}")

        # Check if week already exists
        existing_week_row = self._find_existing_week(ws, week_col, week_label, total_row_num)

        if existing_week_row:
            self.logger.info(f"Week {week_label} already exists at row {existing_week_row}, updating values")
            target_row = existing_week_row
        else:
            # Insert new row above Total
            self.logger.info(f"Inserting new row for week {week_label} above Total")
            target_row = self._insert_new_row(ws, total_row_num, week_col, week_label)

            # Hide oldest row (rolling window)
            if sheet_config.get('rolling_window', True):
                self._hide_oldest_row(ws, week_col, total_row_num)

        # Write data to target row
        self._write_data_to_row(ws, target_row, thisweek_columns, fuel_data, cstore_data)

        self.logger.info(f"Sheet {ws.title} updated successfully")

    def _find_total_row(self, ws, week_col, total_row_label):
        """
        Find the row number of the Total row.

        Args:
            ws: Worksheet
            week_col (str): Column letter for Week Ending (e.g., 'A')
            total_row_label (str): Label to search for (e.g., 'Total')

        Returns:
            int: Row number, or None if not found
        """
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell = row[self._col_letter_to_index(week_col)]
            if cell.value and str(cell.value).strip() == total_row_label:
                return cell.row

        return None

    def _find_existing_week(self, ws, week_col, week_label, total_row_num):
        """
        Check if the week label already exists.

        Args:
            ws: Worksheet
            week_col (str): Column letter
            week_label (str): Week label to search for
            total_row_num (int): Total row number (search above this)

        Returns:
            int: Row number if found, else None
        """
        col_idx = self._col_letter_to_index(week_col)

        for row in ws.iter_rows(min_row=1, max_row=total_row_num - 1):
            cell = row[col_idx]
            if cell.value and str(cell.value).strip() == week_label:
                return cell.row

        return None

    def _insert_new_row(self, ws, total_row_num, week_col, week_label):
        """
        Insert a new row above the Total row and copy formatting.

        Args:
            ws: Worksheet
            total_row_num (int): Total row number
            week_col (str): Week column letter
            week_label (str): Week label to write

        Returns:
            int: Row number of new row
        """
        # Row to insert at (just above Total)
        insert_row_num = total_row_num

        # Get the previous data row (for formatting reference)
        source_row_num = total_row_num - 1

        # Insert row
        ws.insert_rows(insert_row_num, 1)
        self.logger.info(f"Inserted row at {insert_row_num}")

        # Copy formatting from source row to new row
        for col_idx in range(ws.max_column):
            source_cell = ws.cell(row=source_row_num, column=col_idx + 1)
            target_cell = ws.cell(row=insert_row_num, column=col_idx + 1)

            # Copy cell formatting
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Copy row height
        ws.row_dimensions[insert_row_num].height = ws.row_dimensions[source_row_num].height

        # Write week label to week column
        week_cell = ws[f"{week_col}{insert_row_num}"]
        week_cell.value = week_label

        self.logger.info(f"Formatting copied and week label written to row {insert_row_num}")

        return insert_row_num

    def _hide_oldest_row(self, ws, week_col, total_row_num):
        """
        Hide the oldest data row (first data row after headers).

        Args:
            ws: Worksheet
            week_col (str): Week column letter
            total_row_num (int): Total row number
        """
        col_idx = self._col_letter_to_index(week_col)

        # Headers to skip - these are not data rows
        header_keywords = ['week', 'total', 'diesel', 'gas', 'def', 'gallons', 'this week', 'last year']

        # Find first data row (has week label like "2nd June", "9th June")
        for row in ws.iter_rows(min_row=1, max_row=total_row_num):
            cell = row[col_idx]
            if cell.value:
                cell_text = str(cell.value).strip().lower()

                # Skip if empty or matches header keywords
                if not cell_text:
                    continue

                # Check if this is a header row
                is_header = any(keyword in cell_text for keyword in header_keywords)
                if is_header:
                    continue

                # Check if this looks like a week label (has digits)
                has_digits = any(c.isdigit() for c in cell_text)
                if has_digits:
                    # This is the first data row
                    oldest_row_num = cell.row
                    ws.row_dimensions[oldest_row_num].hidden = True
                    self.logger.info(f"Hidden oldest row: {oldest_row_num} ('{cell.value}')")
                    return

        self.logger.warning("Could not find oldest row to hide")

    def _write_data_to_row(self, ws, row_num, column_mapping, fuel_data, cstore_data):
        """
        Write data values to specified columns in a row.

        Args:
            ws: Worksheet
            row_num (int): Row number to write to
            column_mapping (dict): {data_key: column_letter}
            fuel_data (dict): Fuel data
            cstore_data (dict): C-Store data
        """
        # Combine data sources
        all_data = {}
        if fuel_data:
            all_data.update(fuel_data)
        if cstore_data:
            all_data.update(cstore_data)

        for data_key, col_letter in column_mapping.items():
            if data_key in all_data:
                cell = ws[f"{col_letter}{row_num}"]
                cell.value = all_data[data_key]
                self.logger.debug(f"Wrote {data_key}={all_data[data_key]:,.2f} to {col_letter}{row_num}")
            else:
                self.logger.warning(f"Data key '{data_key}' not found in provided data")

    def _col_letter_to_index(self, col_letter):
        """
        Convert column letter to 0-based index.

        Args:
            col_letter (str): Column letter (e.g., 'A', 'B')

        Returns:
            int: 0-based index
        """
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(col_letter) - 1


if __name__ == '__main__':
    # Test updater
    import yaml

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)

    # Test data
    fuel_data = {
        'diesel_gal': 11989.79,
        'regular_gal': 8500.50,
        'def_gal': 1234.56,
        'total_gal': 21724.85
    }

    updater = ExcelUpdater(config)

    try:
        updater.update_tracker('19th Oct', fuel_data)
        print("Update successful!")
    except Exception as e:
        print(f"Update failed: {e}")
