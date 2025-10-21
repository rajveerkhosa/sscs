"""
Historical data backfill script for Weekly Tracker.
Fills empty week rows with historical SSCS data.
"""

import os
import sys
import logging
import yaml
from datetime import datetime, timedelta
from openpyxl import load_workbook
from dotenv import load_dotenv

from sscs_scraper import SSCSScraper
from fuel_aggregator import FuelAggregator
from excel_updater import ExcelUpdater


def parse_week_label_to_date(week_label, year):
    """
    Parse week label like 'Oct 6th' or '22nd Sep' to a date.

    Args:
        week_label (str): Week label from Excel (e.g., 'Oct 6th', '22nd Sep')
        year (int): Year to use for the date

    Returns:
        datetime: Parsed date
    """
    # Remove ordinal suffix (st, nd, rd, th)
    import re

    # Handle both formats: "Oct 6th" and "22nd Sep"
    # Try "Month Day" format first
    match = re.match(r'([A-Za-z]+)\s+(\d+)', week_label)
    if match:
        month_str = match.group(1)
        day = int(match.group(2))
    else:
        # Try "Day Month" format
        match = re.match(r'(\d+)\w+\s+([A-Za-z]+)', week_label)
        if match:
            day = int(match.group(1))
            month_str = match.group(2)
        else:
            raise ValueError(f"Could not parse week label: {week_label}")

    # Parse month abbreviation
    date_str = f"{day} {month_str} {year}"
    date_obj = datetime.strptime(date_str, "%d %b %Y")

    return date_obj


def calculate_week_range(week_ending_date):
    """
    Calculate Monday-Sunday range for a week ending on the given date.

    Args:
        week_ending_date (datetime): The Sunday ending the week

    Returns:
        tuple: (start_datetime, end_datetime)
    """
    # Assume week_ending_date is a Sunday
    # If not Sunday, adjust to the nearest Sunday
    if week_ending_date.weekday() != 6:
        # Find the nearest Sunday
        days_to_sunday = (6 - week_ending_date.weekday()) % 7
        if days_to_sunday == 0:
            days_to_sunday = 7
        week_ending_date = week_ending_date + timedelta(days=days_to_sunday)

    # Set to end of Sunday
    end_datetime = week_ending_date.replace(hour=23, minute=59, second=59, microsecond=0)

    # Calculate Monday (6 days before Sunday)
    start_datetime = (week_ending_date - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)

    return start_datetime, end_datetime


def format_sscs_datetime(dt):
    """Format datetime for SSCS query (YYYYMMDDhhmmss)"""
    return dt.strftime('%Y%m%d%H%M%S')


def get_empty_weeks(config):
    """
    Find all empty weeks in the Excel file.

    Returns:
        list: List of dicts with {row, week_label, start_date, end_date}
    """
    workbook_path = config['excel']['workbook']
    wb = load_workbook(workbook_path)

    empty_weeks = []

    for sheet_config in config['excel']['sheets']:
        sheet_name = sheet_config['name']

        if not sheet_config.get('enabled', True):
            continue

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        week_col = sheet_config['week_col']
        total_row_label = sheet_config['total_row_label']

        # Find Total row
        total_row_num = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(week_col) - 1
            cell = row[col_idx]
            if cell.value and str(cell.value).strip() == total_row_label:
                total_row_num = cell.row
                break

        if total_row_num is None:
            continue

        # Check all rows above Total for empty data
        col_idx = column_index_from_string(week_col) - 1
        data_col = sheet_config['thisweek_columns'].get('diesel_gal', 'B')
        data_col_idx = column_index_from_string(data_col)

        for row in ws.iter_rows(min_row=1, max_row=total_row_num - 1):
            week_cell = row[col_idx]
            data_cell = ws.cell(row=week_cell.row, column=data_col_idx)

            if week_cell.value:
                week_text = str(week_cell.value).strip()

                # Skip headers
                header_keywords = ['week', 'total', 'diesel', 'gas', 'def', 'gallons', 'this week', 'last year']
                is_header = any(keyword in week_text.lower() for keyword in header_keywords)
                if is_header:
                    continue

                # Check if this row has a week label but empty data
                has_digits = any(c.isdigit() for c in week_text)
                if has_digits and (data_cell.value is None or data_cell.value == ''):
                    # Determine year - use 2024 for all these historical weeks
                    # (Sept and early Oct are from 2024)
                    year = 2024

                    try:
                        week_date = parse_week_label_to_date(week_text, year)
                        start_dt, end_dt = calculate_week_range(week_date)

                        empty_weeks.append({
                            'sheet': sheet_name,
                            'row': week_cell.row,
                            'week_label': week_text,
                            'week_date': week_date,
                            'start_date': format_sscs_datetime(start_dt),
                            'end_date': format_sscs_datetime(end_dt),
                            'start_datetime': start_dt,
                            'end_datetime': end_dt
                        })
                    except Exception as e:
                        logging.warning(f"Could not parse week '{week_text}' at row {week_cell.row}: {e}")

    return empty_weeks


def backfill_week(scraper, aggregator, updater, week_info, config):
    """
    Backfill a single week with historical data.

    Args:
        scraper: SSCSScraper instance
        aggregator: FuelAggregator instance
        updater: ExcelUpdater instance
        week_info (dict): Week information
        config (dict): Configuration
    """
    logger = logging.getLogger(__name__)

    logger.info(f"=" * 60)
    logger.info(f"Backfilling: {week_info['week_label']} (Row {week_info['row']})")
    logger.info(f"Date range: {week_info['start_datetime']} to {week_info['end_datetime']}")
    logger.info(f"=" * 60)

    # Collect fuel data
    fuel_data = aggregator.collect_all_gallons(
        week_info['start_date'],
        week_info['end_date']
    )

    logger.info(f"✓ Data collected for {week_info['week_label']}")
    logger.info(f"  Diesel: {fuel_data['diesel_gal']:,.2f} gal")
    logger.info(f"  Regular: {fuel_data['regular_gal']:,.2f} gal")
    logger.info(f"  DEF: {fuel_data['def_gal']:,.2f} gal")
    logger.info(f"  TOTAL: {fuel_data['total_gal']:,.2f} gal")

    # Update Excel
    updater.update_tracker(week_info['week_label'], fuel_data, cstore_data=None)

    logger.info(f"✓ Excel updated for {week_info['week_label']}")


def main():
    """Main backfill routine"""

    # Load environment variables
    load_dotenv()

    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )

    logger = logging.getLogger(__name__)

    print("=" * 60)
    print("SSCS Historical Data Backfill")
    print("=" * 60)

    # Load config
    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)

    # Find empty weeks
    logger.info("Scanning Excel for empty weeks...")
    empty_weeks = get_empty_weeks(config)

    if not empty_weeks:
        logger.info("✓ No empty weeks found - all data is up to date!")
        return

    # Sort by date
    empty_weeks.sort(key=lambda x: x['week_date'])

    logger.info(f"\nFound {len(empty_weeks)} empty week(s):")
    for week in empty_weeks:
        logger.info(f"  - {week['week_label']} (Row {week['row']}): {week['start_datetime'].date()} to {week['end_datetime'].date()}")

    # Confirm with user
    print("\n" + "=" * 60)
    response = input(f"Backfill {len(empty_weeks)} week(s)? [y/N]: ").strip().lower()
    if response != 'y':
        logger.info("Backfill cancelled by user")
        return

    # Initialize scraper and aggregator
    logger.info("\nInitializing SSCS scraper...")
    scraper = SSCSScraper(config, logger)

    # Start browser and login
    headless = os.getenv('HEADLESS', 'true').lower() == 'true'
    scraper.start_browser(headless=headless)
    scraper.login()

    aggregator = FuelAggregator(config, scraper, logger)
    updater = ExcelUpdater(config, logger)

    # Backfill each week
    success_count = 0
    for week in empty_weeks:
        try:
            backfill_week(scraper, aggregator, updater, week, config)
            success_count += 1
        except Exception as e:
            logger.error(f"✗ Failed to backfill {week['week_label']}: {e}")
            import traceback
            traceback.print_exc()

    # Cleanup
    scraper.close()

    # Summary
    print("\n" + "=" * 60)
    print("BACKFILL COMPLETE")
    print("=" * 60)
    logger.info(f"Successfully backfilled {success_count} out of {len(empty_weeks)} weeks")

    if success_count == len(empty_weeks):
        print("✓ All weeks backfilled successfully!")
    else:
        print(f"⚠ {len(empty_weeks) - success_count} week(s) failed - check logs")


if __name__ == '__main__':
    main()
