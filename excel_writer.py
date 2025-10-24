#!/usr/bin/env python3
"""
Excel file generator for SSCS Weekly Report.
Creates Excel workbook with formatted sheets for Fuel, C-Store, and Department Sales.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelReportWriter:
    """Generate formatted Excel reports for SSCS data"""

    def __init__(self, output_path="Weekly_Report.xlsx"):
        """
        Initialize Excel writer.

        Args:
            output_path (str): Path to save the Excel file
        """
        self.output_path = output_path
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

    def _apply_header_style(self, cell):
        """Apply header styling to a cell"""
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_data_style(self, cell, is_number=False):
        """Apply data cell styling"""
        cell.alignment = Alignment(horizontal="right" if is_number else "left", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_total_style(self, cell):
        """Apply total row styling"""
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self._apply_data_style(cell, is_number=True)

    def add_fuel_sheet(self, week_label, week_range, fuel_data, fuel_data_ly, config):
        """
        Add Weekly Gallons sheet.

        Args:
            week_label (str): Week label (e.g., "Sep 29th")
            week_range (str): Week date range
            fuel_data (dict): This week's fuel data
            fuel_data_ly (dict): Last year's fuel data
            config (dict): Config with prefix lists
        """
        ws = self.wb.create_sheet("Weekly Gallons 25")

        # Title
        ws['A1'] = "WEEKLY GALLONS 25"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Week: {week_range}"
        ws['A3'] = f"Excel Row Label: {week_label}"

        # Headers
        headers = ["Fuel Type", "THIS WEEK", "LAST YEAR", "CHANGE", "% CHANGE"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            self._apply_header_style(cell)

        row = 6

        # Diesel
        diesel_change = fuel_data['diesel_gal'] - fuel_data_ly['diesel_gal']
        diesel_pct = (diesel_change / fuel_data_ly['diesel_gal'] * 100) if fuel_data_ly['diesel_gal'] != 0 else 0

        ws[f'A{row}'] = "Diesel (B/C)"
        ws[f'B{row}'] = fuel_data['diesel_gal']
        ws[f'C{row}'] = fuel_data_ly['diesel_gal']
        ws[f'D{row}'] = diesel_change
        ws[f'E{row}'] = diesel_pct
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '+#,##0.00;-#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
        row += 1

        # Diesel prefixes
        for prefix in config['fuel']['diesel_prefixes']:
            tw = fuel_data['prefix_details'].get(prefix, 0)
            ly = fuel_data_ly['prefix_details'].get(prefix, 0)
            change = tw - ly
            pct = (change / ly * 100) if ly != 0 else 0

            ws[f'A{row}'] = f"  Prefix {prefix}"
            ws[f'B{row}'] = tw
            ws[f'C{row}'] = ly
            ws[f'D{row}'] = change
            ws[f'E{row}'] = pct
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'C{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].number_format = '+#,##0.00;-#,##0.00'
            ws[f'E{row}'].number_format = '0.00"%"'

            for col in range(1, 6):
                self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
            row += 1

        # Regular
        regular_change = fuel_data['regular_gal'] - fuel_data_ly['regular_gal']
        regular_pct = (regular_change / fuel_data_ly['regular_gal'] * 100) if fuel_data_ly['regular_gal'] != 0 else 0

        ws[f'A{row}'] = "Regular (E/F)"
        ws[f'B{row}'] = fuel_data['regular_gal']
        ws[f'C{row}'] = fuel_data_ly['regular_gal']
        ws[f'D{row}'] = regular_change
        ws[f'E{row}'] = regular_pct
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '+#,##0.00;-#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
        row += 1

        # Regular prefixes
        for prefix in config['fuel']['regular_prefixes']:
            tw = fuel_data['prefix_details'].get(prefix, 0)
            ly = fuel_data_ly['prefix_details'].get(prefix, 0)
            change = tw - ly
            pct = (change / ly * 100) if ly != 0 else 0

            ws[f'A{row}'] = f"  Prefix {prefix}"
            ws[f'B{row}'] = tw
            ws[f'C{row}'] = ly
            ws[f'D{row}'] = change
            ws[f'E{row}'] = pct
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'C{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].number_format = '+#,##0.00;-#,##0.00'
            ws[f'E{row}'].number_format = '0.00"%"'

            for col in range(1, 6):
                self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
            row += 1

        # DEF
        def_change = fuel_data['def_gal'] - fuel_data_ly['def_gal']
        def_pct = (def_change / fuel_data_ly['def_gal'] * 100) if fuel_data_ly['def_gal'] != 0 else 0

        ws[f'A{row}'] = "DEF (H/I)"
        ws[f'B{row}'] = fuel_data['def_gal']
        ws[f'C{row}'] = fuel_data_ly['def_gal']
        ws[f'D{row}'] = def_change
        ws[f'E{row}'] = def_pct
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '+#,##0.00;-#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
        row += 1

        # DEF prefix
        for prefix in config['fuel']['def_prefixes']:
            tw = fuel_data['prefix_details'].get(prefix, 0)
            ly = fuel_data_ly['prefix_details'].get(prefix, 0)
            change = tw - ly
            pct = (change / ly * 100) if ly != 0 else 0

            ws[f'A{row}'] = f"  Prefix {prefix}"
            ws[f'B{row}'] = tw
            ws[f'C{row}'] = ly
            ws[f'D{row}'] = change
            ws[f'E{row}'] = pct
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'C{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].number_format = '+#,##0.00;-#,##0.00'
            ws[f'E{row}'].number_format = '0.00"%"'

            for col in range(1, 6):
                self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
            row += 1

        # Total row
        total_change = fuel_data['total_gal'] - fuel_data_ly['total_gal']
        total_pct = (total_change / fuel_data_ly['total_gal'] * 100) if fuel_data_ly['total_gal'] != 0 else 0

        ws[f'A{row}'] = "TOTAL (K/L)"
        ws[f'B{row}'] = fuel_data['total_gal']
        ws[f'C{row}'] = fuel_data_ly['total_gal']
        ws[f'D{row}'] = total_change
        ws[f'E{row}'] = total_pct
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '+#,##0.00;-#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_total_style(ws.cell(row=row, column=col))

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

    def add_cstore_sheet(self, week_label, week_range, cstore_data, cstore_data_ly):
        """
        Add C Store Sales sheet.

        Args:
            week_label (str): Week label
            week_range (str): Week date range
            cstore_data (dict): This week's C-Store data
            cstore_data_ly (dict): Last year's C-Store data
        """
        ws = self.wb.create_sheet("C Store Sales 25")

        # Title
        ws['A1'] = "C STORE SALES 25"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Week: {week_range}"
        ws['A3'] = f"Excel Row Label: {week_label}"

        # Headers
        headers = ["Category", "THIS WEEK", "LAST YEAR", "CHANGE", "% CHANGE"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            self._apply_header_style(cell)

        row = 6

        # Total C-Store
        total_change = cstore_data['total_cstore_sales'] - cstore_data_ly['total_cstore_sales']
        total_pct = (total_change / cstore_data_ly['total_cstore_sales'] * 100) if cstore_data_ly['total_cstore_sales'] != 0 else 0

        ws[f'A{row}'] = "Total C-Store (B/C)"
        ws[f'B{row}'] = cstore_data['total_cstore_sales']
        ws[f'C{row}'] = cstore_data_ly['total_cstore_sales']
        ws[f'D{row}'] = total_change
        ws[f'E{row}'] = total_pct
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].number_format = '+$#,##0.00;-$#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
        row += 1

        # Lottery
        lottery_change = cstore_data['lottery_sales'] - cstore_data_ly['lottery_sales']
        lottery_pct = (lottery_change / cstore_data_ly['lottery_sales'] * 100) if cstore_data_ly['lottery_sales'] != 0 else 0

        ws[f'A{row}'] = "Lottery (E/F)"
        ws[f'B{row}'] = cstore_data['lottery_sales']
        ws[f'C{row}'] = cstore_data_ly['lottery_sales']
        ws[f'D{row}'] = lottery_change
        ws[f'E{row}'] = lottery_pct
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].number_format = '+$#,##0.00;-$#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
        row += 1

        # Lottery departments
        for dept_num in ['27', '43', '72']:
            tw = cstore_data.get(f'dept_{dept_num}', 0)
            ly = cstore_data_ly.get(f'dept_{dept_num}', 0)
            change = tw - ly
            pct = (change / ly * 100) if ly != 0 else 0

            ws[f'A{row}'] = f"  Dept {dept_num}"
            ws[f'B{row}'] = tw
            ws[f'C{row}'] = ly
            ws[f'D{row}'] = change
            ws[f'E{row}'] = pct
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'C{row}'].number_format = '$#,##0.00'
            ws[f'D{row}'].number_format = '+$#,##0.00;-$#,##0.00'
            ws[f'E{row}'].number_format = '0.00"%"'

            for col in range(1, 6):
                self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
            row += 1

        # Scale
        scale_change = cstore_data['scale_sales'] - cstore_data_ly['scale_sales']
        scale_pct = (scale_change / cstore_data_ly['scale_sales'] * 100) if cstore_data_ly['scale_sales'] != 0 else 0

        ws[f'A{row}'] = "Scale (H/I)"
        ws[f'B{row}'] = cstore_data['scale_sales']
        ws[f'C{row}'] = cstore_data_ly['scale_sales']
        ws[f'D{row}'] = scale_change
        ws[f'E{row}'] = scale_pct
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].number_format = '+$#,##0.00;-$#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
        row += 1

        # Dept 88
        tw = cstore_data.get('dept_88', 0)
        ly = cstore_data_ly.get('dept_88', 0)
        change = tw - ly
        pct = (change / ly * 100) if ly != 0 else 0

        ws[f'A{row}'] = "  Dept 88"
        ws[f'B{row}'] = tw
        ws[f'C{row}'] = ly
        ws[f'D{row}'] = change
        ws[f'E{row}'] = pct
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].number_format = '+$#,##0.00;-$#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_data_style(ws.cell(row=row, column=col), is_number=(col > 1))
        row += 1

        # Other Sales
        other_change = cstore_data['other_sales'] - cstore_data_ly['other_sales']
        other_pct = (other_change / cstore_data_ly['other_sales'] * 100) if cstore_data_ly['other_sales'] != 0 else 0

        ws[f'A{row}'] = "Other Sales (K/L)"
        ws[f'B{row}'] = cstore_data['other_sales']
        ws[f'C{row}'] = cstore_data_ly['other_sales']
        ws[f'D{row}'] = other_change
        ws[f'E{row}'] = other_pct
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].number_format = '+$#,##0.00;-$#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'

        for col in range(1, 6):
            self._apply_total_style(ws.cell(row=row, column=col))

        # Adjust column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

    def add_department_sheet(self, week_label, week_range, dept_sales, dept_names):
        """
        Add Department Sales sheet.

        Args:
            week_label (str): Week label
            week_range (str): Week date range
            dept_sales (dict): Department sales {dept_num: sales}
            dept_names (dict): Department names {dept_num: name}
        """
        ws = self.wb.create_sheet("Weekly Department Sales")

        # Title
        ws['A1'] = "WEEKLY DEPARTMENT SALES"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Week: {week_range}"
        ws['A3'] = f"Column Header: {week_label}"

        # Headers
        headers = ["Dept #", "Department Name", "Sales"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            self._apply_header_style(cell)

        row = 6
        total_sales = 0

        for dept_num in sorted(dept_sales.keys(), key=lambda x: int(x)):
            sales = dept_sales[dept_num]
            total_sales += sales

            ws[f'A{row}'] = int(dept_num)
            ws[f'B{row}'] = dept_names.get(dept_num, "")
            ws[f'C{row}'] = sales
            ws[f'C{row}'].number_format = '$#,##0.00'

            for col in range(1, 4):
                self._apply_data_style(ws.cell(row=row, column=col), is_number=(col == 3))
            row += 1

        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = ""
        ws[f'C{row}'] = total_sales
        ws[f'C{row}'].number_format = '$#,##0.00'

        for col in range(1, 4):
            self._apply_total_style(ws.cell(row=row, column=col))

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15

    def save(self):
        """Save the workbook to file"""
        self.wb.save(self.output_path)
        return self.output_path


if __name__ == '__main__':
    # Test the Excel writer
    import yaml

    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)

    # Sample data
    fuel_data = {
        'diesel_gal': 15860.75,
        'regular_gal': 9764.31,
        'def_gal': 425.99,
        'total_gal': 26051.05,
        'prefix_details': {
            '050': 11797.27,
            '019': 4063.48,
            '001': 7788.67,
            '002': 1058.80,
            '003': 916.84,
            '062': 425.99
        }
    }

    fuel_data_ly = {
        'diesel_gal': 14485.53,
        'regular_gal': 10058.25,
        'def_gal': 325.30,
        'total_gal': 24869.08,
        'prefix_details': {
            '050': 10701.58,
            '019': 3783.95,
            '001': 8169.34,
            '002': 928.65,
            '003': 960.26,
            '062': 325.30
        }
    }

    cstore_data = {
        'total_cstore_sales': 47710.24,
        'lottery_sales': 4314.00,
        'scale_sales': 2290.75,
        'other_sales': 41105.49,
        'dept_27': 2670.00,
        'dept_43': 591.00,
        'dept_72': 1053.00,
        'dept_88': 2290.75
    }

    cstore_data_ly = {
        'total_cstore_sales': 47734.96,
        'lottery_sales': 4349.00,
        'scale_sales': 1330.99,
        'other_sales': 42054.97,
        'dept_27': 3764.00,
        'dept_43': 0.00,
        'dept_72': 585.00,
        'dept_88': 1330.99
    }

    dept_sales = {'79': 4231.59, '55': 0.00, '59': 3035.81}
    dept_names = {'79': 'Krispy Krunchy', '55': 'Beer High Desert', '59': 'Core mark no tax'}

    writer = ExcelReportWriter("test_report.xlsx")
    writer.add_fuel_sheet("Sep 29th", "2025-09-22 to 2025-09-28", fuel_data, fuel_data_ly, config)
    writer.add_cstore_sheet("Sep 29th", "2025-09-22 to 2025-09-28", cstore_data, cstore_data_ly)
    writer.add_department_sheet("29th Sep", "2025-09-22 to 2025-09-28", dept_sales, dept_names)
    path = writer.save()
    print(f"✓ Test Excel file created: {path}")
